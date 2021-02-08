import openpyxl

def getRowCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return  (sheet.max_row)

def getcolumnCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return  (sheet.max_column)

def readData(file,SheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(row=rownum, column=columnnum).value

def writeData(file,SheetName,rownum,columnnum,Data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(row=rownum, column=columnnum).value= Data
    workbook.save(file)